"""XLSX parser — extracts visible worksheet values as bounded Markdown tables."""

from __future__ import annotations

import io
from typing import Optional

from agent.modules.knowledge.schemas import NormalizedDocument


def parse(
    raw: bytes,
    *,
    title: str = "",
    author: str = "",
    source_type: str = "project_doc",
    scope: str = "workspace",
    language: str = "zh",
    metadata: Optional[dict] = None,
) -> NormalizedDocument:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        return _empty(title, author, source_type, scope, language, metadata, f"xlsx_parser_unavailable: {exc!r}")
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        return _empty(title, author, source_type, scope, language, metadata, f"xlsx_open_failed: {exc!r}")

    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"## {sheet.title}")
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value).replace("|", "\\|").replace("\n", " ") for value in row]
            if any(values):
                rows.append(values)
            if len(rows) >= 500:
                break
        if not rows:
            lines.append("（空工作表）")
            continue
        width = max(len(row) for row in rows)
        rows = [row + [""] * (width - len(row)) for row in rows]
        header = rows[0]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join("---" for _ in header) + " |")
        lines.extend("| " + " | ".join(row) + " |" for row in rows[1:])
        if sheet.max_row > len(rows):
            lines.append("（该工作表仅提取前 500 个非空行）")
        lines.append("")
    workbook.close()
    return NormalizedDocument(
        title=title,
        author=author,
        source_type=source_type,
        scope=scope,
        language=language,
        format="xlsx",
        normalized_markdown="\n".join(lines).strip(),
        metadata={**(metadata or {}), "format_hint": "xlsx"},
        warnings=[],
    )


def _empty(title, author, source_type, scope, language, metadata, warning):
    return NormalizedDocument(
        title=title, author=author, source_type=source_type, scope=scope,
        language=language, format="xlsx", normalized_markdown="",
        metadata=metadata or {}, warnings=[warning],
    )
